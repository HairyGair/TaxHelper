"""
Compliance & Audit Reports Generator
HMRC-ready documentation and export utilities
"""

import os
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path
import io

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, Image, KeepTogether
)
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import pandas as pd
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_

from models import (
    Transaction, AuditLog, TransactionType
)


# ============================================================================
# CONFIGURATION
# ============================================================================

REPORTS_DIR = Path("/Users/anthony/Tax Helper/reports")
REPORTS_DIR.mkdir(exist_ok=True)

# Tax year boundaries (UK: 6 April to 5 April)
def get_tax_year_dates(tax_year: str) -> Tuple[datetime, datetime]:
    """
    Convert tax year string to date range.
    Example: "2024/25" -> (2024-04-06, 2025-04-05)
    """
    start_year = int(tax_year.split('/')[0])
    end_year = start_year + 1

    start_date = datetime(start_year, 4, 6, 0, 0, 0)
    end_date = datetime(end_year, 4, 5, 23, 59, 59)

    return start_date, end_date


def get_current_tax_year() -> str:
    """Get current tax year string based on today's date."""
    today = datetime.now()

    if today.month < 4 or (today.month == 4 and today.day < 6):
        # Before 6 April - still in previous tax year
        start_year = today.year - 1
    else:
        # After 6 April - new tax year
        start_year = today.year

    end_year = start_year + 1
    return f"{start_year}/{str(end_year)[2:]}"


# ============================================================================
# 1. AUDIT TRAIL REPORT
# ============================================================================

def generate_audit_trail_report(
    session: Session,
    tax_year_start: datetime,
    tax_year_end: datetime,
    format: str = 'PDF'
) -> str:
    """
    Generate complete audit trail report for HMRC compliance.

    Args:
        session: Database session
        tax_year_start: Start of tax year
        tax_year_end: End of tax year
        format: Output format ('PDF', 'CSV', 'Excel')

    Returns:
        Path to generated report file
    """
    # Gather data
    transactions = session.query(Transaction).filter(
        and_(
            Transaction.date >= tax_year_start,
            Transaction.date <= tax_year_end
        )
    ).all()

    audit_logs = session.query(AuditLog).filter(
        and_(
            AuditLog.timestamp >= tax_year_start,
            AuditLog.timestamp <= tax_year_end
        )
    ).order_by(AuditLog.timestamp.desc()).all()

    income_count = sum(1 for t in transactions if t.type == TransactionType.INCOME)
    expense_count = sum(1 for t in transactions if t.type == TransactionType.EXPENSE)

    # Calculate summary stats
    total_changes = len(audit_logs)

    # Generate report based on format
    tax_year_str = f"{tax_year_start.year}/{str(tax_year_end.year)[2:]}"

    if format.upper() == 'PDF':
        return _generate_audit_trail_pdf(
            transactions, audit_logs, income_count, expense_count,
            total_changes, tax_year_str, tax_year_start, tax_year_end
        )
    elif format.upper() == 'CSV':
        return _generate_audit_trail_csv(
            audit_logs, tax_year_str
        )
    elif format.upper() == 'EXCEL':
        return _generate_audit_trail_excel(
            transactions, audit_logs, income_count, expense_count,
            total_changes, tax_year_str
        )
    else:
        raise ValueError(f"Unsupported format: {format}")


def _generate_audit_trail_pdf(
    transactions, audit_logs, income_count, expense_count,
    total_changes, tax_year_str, tax_year_start, tax_year_end
) -> str:
    """Generate audit trail PDF report."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"audit_trail_{tax_year_str.replace('/', '_')}_{timestamp}.pdf"

    doc = SimpleDocTemplate(
        str(filename),
        pagesize=A4,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )

    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=12,
        borderPadding=5,
        backColor=colors.HexColor('#ecf0f1')
    )

    # Cover page
    story.append(Paragraph("AUDIT TRAIL REPORT", title_style))
    story.append(Spacer(1, 0.3*inch))

    cover_info = f"""
    <para alignment="center">
    <b>Tax Year {tax_year_str}</b><br/>
    ({tax_year_start.strftime('%d %B %Y')} - {tax_year_end.strftime('%d %B %Y')})<br/>
    <br/>
    Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}<br/>
    <br/>
    Tax Helper v3.0
    </para>
    """
    story.append(Paragraph(cover_info, styles['Normal']))
    story.append(PageBreak())

    # Summary section
    story.append(Paragraph("SUMMARY", heading_style))
    story.append(Spacer(1, 0.2*inch))

    summary_data = [
        ['Metric', 'Count'],
        ['Total Transactions', str(len(transactions))],
        ['Income Records', str(income_count)],
        ['Expense Records', str(expense_count)],
        ['Total Changes Logged', str(total_changes)],
        ['Audit Log Entries', str(len(audit_logs))]
    ]

    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))

    story.append(summary_table)
    story.append(PageBreak())

    # Detailed log section
    story.append(Paragraph("DETAILED AUDIT LOG", heading_style))
    story.append(Spacer(1, 0.2*inch))

    for log in audit_logs[:100]:  # Limit to first 100 for PDF
        log_entry = f"""
        <b>{log.timestamp.strftime('%Y-%m-%d %H:%M:%S')}</b> | Transaction #{log.transaction_id or 'N/A'}<br/>
        <b>Action:</b> {log.action}<br/>
        <b>Before:</b> {log.before_value or 'null'}<br/>
        <b>After:</b> {log.after_value or 'null'}<br/>
        <b>Summary:</b> {log.change_summary or 'No summary'}<br/>
        """
        story.append(Paragraph(log_entry, styles['Normal']))
        story.append(Spacer(1, 0.15*inch))

    if len(audit_logs) > 100:
        story.append(Paragraph(
            f"<i>Note: Showing first 100 of {len(audit_logs)} audit entries. "
            "See Excel/CSV export for complete log.</i>",
            styles['Italic']
        ))

    story.append(PageBreak())

    # Methodology section
    story.append(Paragraph("METHODOLOGY", heading_style))
    story.append(Spacer(1, 0.2*inch))

    methodology = """
    This audit trail was generated using Tax Helper v3.0, a comprehensive tax record
    management system designed for HMRC compliance.<br/>
    <br/>
    <b>Data Collection:</b><br/>
    All changes to transaction records are logged automatically with precise timestamps,
    before/after values, and descriptive change summaries. No manual intervention
    affects the audit trail.<br/>
    <br/>
    <b>Data Integrity:</b><br/>
    The audit log is immutable once written. All entries are stored in a SQLite database
    with transaction integrity guarantees.<br/>
    <br/>
    <b>Categorization Methods:</b><br/>
    - Merchant matching against known patterns<br/>
    - Rule-based classification using user-defined rules<br/>
    - Pattern learning from historical categorizations<br/>
    - Manual categorization with confidence tracking<br/>
    <br/>
    <b>Compliance:</b><br/>
    This report meets HMRC Making Tax Digital (MTD) requirements for record keeping
    and audit trail documentation.
    """
    story.append(Paragraph(methodology, styles['Normal']))

    # Build PDF
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)

    return str(filename)


def _generate_audit_trail_csv(audit_logs, tax_year_str) -> str:
    """Generate audit trail CSV export."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"audit_trail_{tax_year_str.replace('/', '_')}_{timestamp}.csv"

    data = []
    for log in audit_logs:
        data.append({
            'Timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'Transaction ID': log.transaction_id or '',
            'Action': log.action,
            'Before Value': log.before_value or '',
            'After Value': log.after_value or '',
            'Change Summary': log.change_summary or ''
        })

    df = pd.DataFrame(data)
    df.to_csv(filename, index=False)

    return str(filename)


def _generate_audit_trail_excel(
    transactions, audit_logs, income_count, expense_count,
    total_changes, tax_year_str
) -> str:
    """Generate audit trail Excel workbook."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"audit_trail_{tax_year_str.replace('/', '_')}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _format_excel_summary_sheet(
        ws_summary, transactions, income_count, expense_count,
        total_changes, len(audit_logs), tax_year_str
    )

    # Audit log sheet
    ws_log = wb.create_sheet("Audit Log")
    _format_excel_audit_log_sheet(ws_log, audit_logs)

    wb.save(filename)
    return str(filename)


# ============================================================================
# 2. RECEIPT SUMMARY REPORT
# ============================================================================

def generate_receipt_summary(
    session: Session,
    tax_year_start: datetime,
    tax_year_end: datetime,
    format: str = 'PDF'
) -> str:
    """
    Generate receipt summary report organized by category.

    Args:
        session: Database session
        tax_year_start: Start of tax year
        tax_year_end: End of tax year
        format: Output format ('PDF', 'CSV', 'Excel')

    Returns:
        Path to generated report file
    """
    # Get all transactions with receipts in tax year
    transactions = session.query(Transaction).join(Receipt).filter(
        and_(
            Transaction.date >= tax_year_start,
            Transaction.date <= tax_year_end,
            Transaction.type == TransactionType.EXPENSE
        )
    ).order_by(Transaction.category, Transaction.date.desc()).all()

    # Group by category
    receipts_by_category = {}
    total_amount = 0
    total_receipts = 0

    for trans in transactions:
        if not trans.receipts:
            continue

        category = trans.category or "Uncategorized"
        if category not in receipts_by_category:
            receipts_by_category[category] = {
                'transactions': [],
                'total': 0,
                'count': 0
            }

        receipts_by_category[category]['transactions'].append(trans)
        receipts_by_category[category]['total'] += abs(trans.amount)
        receipts_by_category[category]['count'] += len(trans.receipts)

        total_amount += abs(trans.amount)
        total_receipts += len(trans.receipts)

    tax_year_str = f"{tax_year_start.year}/{str(tax_year_end.year)[2:]}"

    if format.upper() == 'PDF':
        return _generate_receipt_summary_pdf(
            receipts_by_category, total_receipts, total_amount,
            tax_year_str, tax_year_start, tax_year_end
        )
    elif format.upper() == 'CSV':
        return _generate_receipt_summary_csv(
            receipts_by_category, tax_year_str
        )
    elif format.upper() == 'EXCEL':
        return _generate_receipt_summary_excel(
            receipts_by_category, total_receipts, total_amount, tax_year_str
        )
    else:
        raise ValueError(f"Unsupported format: {format}")


def _generate_receipt_summary_pdf(
    receipts_by_category, total_receipts, total_amount,
    tax_year_str, tax_year_start, tax_year_end
) -> str:
    """Generate receipt summary PDF report."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"receipt_summary_{tax_year_str.replace('/', '_')}_{timestamp}.pdf"

    doc = SimpleDocTemplate(
        str(filename),
        pagesize=A4,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=12,
        backColor=colors.HexColor('#ecf0f1')
    )

    category_style = ParagraphStyle(
        'CategoryHeading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#16a085'),
        spaceAfter=8,
        spaceBefore=8
    )

    # Cover page
    story.append(Paragraph("RECEIPT SUMMARY REPORT", title_style))
    story.append(Spacer(1, 0.3*inch))

    cover_info = f"""
    <para alignment="center">
    <b>Tax Year {tax_year_str}</b><br/>
    ({tax_year_start.strftime('%d %B %Y')} - {tax_year_end.strftime('%d %B %Y')})<br/>
    <br/>
    Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}<br/>
    </para>
    """
    story.append(Paragraph(cover_info, styles['Normal']))
    story.append(PageBreak())

    # Summary totals
    story.append(Paragraph("SUMMARY", heading_style))
    story.append(Spacer(1, 0.2*inch))

    summary = f"""
    <b>Total Receipts:</b> {total_receipts}<br/>
    <b>Total Amount:</b> £{total_amount:,.2f}<br/>
    <b>Categories:</b> {len(receipts_by_category)}<br/>
    """
    story.append(Paragraph(summary, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))

    # Category breakdowns
    for category in sorted(receipts_by_category.keys()):
        data = receipts_by_category[category]

        category_header = f"{category.upper()} - Total: £{data['total']:,.2f} ({data['count']} receipts)"
        story.append(Paragraph(category_header, category_style))
        story.append(Spacer(1, 0.1*inch))

        # Receipt table
        receipt_data = [['Date', 'Merchant', 'Amount', 'Receipt File']]

        for trans in data['transactions'][:20]:  # Limit per category
            for receipt in trans.receipts:
                receipt_data.append([
                    trans.date.strftime('%d/%m/%Y'),
                    trans.merchant or 'Unknown',
                    f"£{abs(trans.amount):,.2f}",
                    Path(receipt.file_path).name if receipt.file_path else 'N/A'
                ])

        if receipt_data:
            receipt_table = Table(receipt_data, colWidths=[1*inch, 2*inch, 1.2*inch, 2.3*inch])
            receipt_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a085')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))

            story.append(receipt_table)

        if len(data['transactions']) > 20:
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                f"<i>Showing first 20 of {len(data['transactions'])} receipts in this category.</i>",
                styles['Italic']
            ))

        story.append(Spacer(1, 0.3*inch))

    # Build PDF
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)

    return str(filename)


def _generate_receipt_summary_csv(receipts_by_category, tax_year_str) -> str:
    """Generate receipt summary CSV export."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"receipt_summary_{tax_year_str.replace('/', '_')}_{timestamp}.csv"

    data = []
    for category, cat_data in sorted(receipts_by_category.items()):
        for trans in cat_data['transactions']:
            for receipt in trans.receipts:
                data.append({
                    'Category': category,
                    'Date': trans.date.strftime('%Y-%m-%d'),
                    'Merchant': trans.merchant or '',
                    'Amount': f"{abs(trans.amount):.2f}",
                    'Receipt File': Path(receipt.file_path).name if receipt.file_path else '',
                    'File Path': receipt.file_path or ''
                })

    df = pd.DataFrame(data)
    df.to_csv(filename, index=False)

    return str(filename)


def _generate_receipt_summary_excel(
    receipts_by_category, total_receipts, total_amount, tax_year_str
) -> str:
    """Generate receipt summary Excel workbook."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"receipt_summary_{tax_year_str.replace('/', '_')}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipts by Category"

    # Headers
    headers = ['Category', 'Date', 'Merchant', 'Amount', 'Receipt File', 'File Path']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="16a085", end_color="16a085", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    row_num = 2
    for category in sorted(receipts_by_category.keys()):
        cat_data = receipts_by_category[category]

        for trans in cat_data['transactions']:
            for receipt in trans.receipts:
                ws.append([
                    category,
                    trans.date.strftime('%d/%m/%Y'),
                    trans.merchant or '',
                    abs(trans.amount),
                    Path(receipt.file_path).name if receipt.file_path else '',
                    receipt.file_path or ''
                ])

                # Format amount as currency
                ws.cell(row=row_num, column=4).number_format = '£#,##0.00'
                row_num += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(['Metric', 'Value'])
    ws_summary.append(['Total Receipts', total_receipts])
    ws_summary.append(['Total Amount', total_amount])
    ws_summary.append(['Categories', len(receipts_by_category)])

    # Format summary
    ws_summary['A1'].font = Font(bold=True)
    ws_summary['B1'].font = Font(bold=True)
    ws_summary['B2'].number_format = '0'
    ws_summary['B3'].number_format = '£#,##0.00'
    ws_summary['B4'].number_format = '0'

    wb.save(filename)
    return str(filename)


# ============================================================================
# 3. CATEGORIZATION REPORT
# ============================================================================

def generate_categorization_report(
    session: Session,
    tax_year_start: datetime,
    tax_year_end: datetime
) -> str:
    """
    Generate categorization methodology report.

    Shows how transactions were classified and confidence levels.
    """
    transactions = session.query(Transaction).filter(
        and_(
            Transaction.date >= tax_year_start,
            Transaction.date <= tax_year_end
        )
    ).all()

    # Analyze categorization methods
    auto_merchant = sum(1 for t in transactions if t.categorization_method == 'merchant_match')
    auto_rule = sum(1 for t in transactions if t.categorization_method == 'rule_match')
    auto_pattern = sum(1 for t in transactions if t.categorization_method == 'pattern_learning')
    manual = sum(1 for t in transactions if t.categorization_method == 'manual')

    total_auto = auto_merchant + auto_rule + auto_pattern

    # Confidence breakdown
    high_conf = sum(1 for t in transactions if t.confidence_score >= 70)
    med_conf = sum(1 for t in transactions if 40 <= t.confidence_score < 70)
    low_conf = sum(1 for t in transactions if 10 <= t.confidence_score < 40)
    no_conf = sum(1 for t in transactions if t.confidence_score < 10)

    # Top merchants
    merchant_counts = {}
    for trans in transactions:
        if trans.merchant:
            merchant_counts[trans.merchant] = merchant_counts.get(trans.merchant, 0) + 1

    top_merchants = sorted(merchant_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    # Category breakdown
    category_stats = {}
    for trans in transactions:
        if trans.category:
            if trans.category not in category_stats:
                category_stats[trans.category] = {
                    'count': 0,
                    'total_confidence': 0
                }
            category_stats[trans.category]['count'] += 1
            category_stats[trans.category]['total_confidence'] += trans.confidence_score or 0

    # Calculate averages
    for cat in category_stats:
        count = category_stats[cat]['count']
        category_stats[cat]['avg_confidence'] = (
            category_stats[cat]['total_confidence'] / count if count > 0 else 0
        )

    tax_year_str = f"{tax_year_start.year}/{str(tax_year_end.year)[2:]}"

    return _generate_categorization_pdf(
        len(transactions), total_auto, auto_merchant, auto_rule, auto_pattern,
        manual, high_conf, med_conf, low_conf, no_conf, top_merchants,
        category_stats, tax_year_str, tax_year_start, tax_year_end
    )


def _generate_categorization_pdf(
    total_trans, total_auto, auto_merchant, auto_rule, auto_pattern,
    manual, high_conf, med_conf, low_conf, no_conf, top_merchants,
    category_stats, tax_year_str, tax_year_start, tax_year_end
) -> str:
    """Generate categorization report PDF."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"categorization_report_{tax_year_str.replace('/', '_')}_{timestamp}.pdf"

    doc = SimpleDocTemplate(
        str(filename),
        pagesize=A4,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=12,
        backColor=colors.HexColor('#ecf0f1')
    )

    # Cover page
    story.append(Paragraph("CATEGORIZATION REPORT", title_style))
    story.append(Spacer(1, 0.3*inch))

    cover_info = f"""
    <para alignment="center">
    <b>Tax Year {tax_year_str}</b><br/>
    ({tax_year_start.strftime('%d %B %Y')} - {tax_year_end.strftime('%d %B %Y')})<br/>
    <br/>
    Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}<br/>
    </para>
    """
    story.append(Paragraph(cover_info, styles['Normal']))
    story.append(PageBreak())

    # Classification methods
    story.append(Paragraph("CLASSIFICATION METHODS", heading_style))
    story.append(Spacer(1, 0.2*inch))

    auto_pct = (total_auto / total_trans * 100) if total_trans > 0 else 0
    manual_pct = (manual / total_trans * 100) if total_trans > 0 else 0

    methods_info = f"""
    <b>Total Transactions:</b> {total_trans:,}<br/>
    <br/>
    <b>Auto-Categorized:</b> {total_auto:,} ({auto_pct:.1f}%)<br/>
    &nbsp;&nbsp;&nbsp;&nbsp;Merchant Match: {auto_merchant:,} ({auto_merchant/total_trans*100 if total_trans > 0 else 0:.1f}%)<br/>
    &nbsp;&nbsp;&nbsp;&nbsp;Rule Match: {auto_rule:,} ({auto_rule/total_trans*100 if total_trans > 0 else 0:.1f}%)<br/>
    &nbsp;&nbsp;&nbsp;&nbsp;Pattern Learning: {auto_pattern:,} ({auto_pattern/total_trans*100 if total_trans > 0 else 0:.1f}%)<br/>
    <br/>
    <b>Manual Categorization:</b> {manual:,} ({manual_pct:.1f}%)<br/>
    """
    story.append(Paragraph(methods_info, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))

    # Confidence breakdown
    story.append(Paragraph("CONFIDENCE BREAKDOWN", heading_style))
    story.append(Spacer(1, 0.2*inch))

    conf_data = [
        ['Confidence Level', 'Count', 'Percentage'],
        [
            'High (70-100%)',
            f"{high_conf:,}",
            f"{high_conf/total_trans*100 if total_trans > 0 else 0:.1f}%"
        ],
        [
            'Medium (40-69%)',
            f"{med_conf:,}",
            f"{med_conf/total_trans*100 if total_trans > 0 else 0:.1f}%"
        ],
        [
            'Low (10-39%)',
            f"{low_conf:,}",
            f"{low_conf/total_trans*100 if total_trans > 0 else 0:.1f}%"
        ],
        [
            'None (0-9%)',
            f"{no_conf:,}",
            f"{no_conf/total_trans*100 if total_trans > 0 else 0:.1f}%"
        ]
    ]

    conf_table = Table(conf_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    conf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))

    story.append(conf_table)
    story.append(Spacer(1, 0.3*inch))

    # Top merchants
    story.append(Paragraph("TOP MERCHANTS", heading_style))
    story.append(Spacer(1, 0.2*inch))

    merchant_data = [['Rank', 'Merchant', 'Transactions']]
    for idx, (merchant, count) in enumerate(top_merchants, 1):
        merchant_data.append([str(idx), merchant, f"{count:,}"])

    merchant_table = Table(merchant_data, colWidths=[0.75*inch, 3.5*inch, 1.75*inch])
    merchant_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a085')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))

    story.append(merchant_table)
    story.append(PageBreak())

    # By category
    story.append(Paragraph("CATEGORIZATION BY CATEGORY", heading_style))
    story.append(Spacer(1, 0.2*inch))

    cat_data = [['Category', 'Count', 'Percentage', 'Avg Confidence']]
    for category in sorted(category_stats.keys()):
        stats = category_stats[category]
        cat_data.append([
            category,
            f"{stats['count']:,}",
            f"{stats['count']/total_trans*100 if total_trans > 0 else 0:.1f}%",
            f"{stats['avg_confidence']:.0f}%"
        ])

    cat_table = Table(cat_data, colWidths=[2.5*inch, 1.25*inch, 1.25*inch, 1.25*inch])
    cat_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e67e22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))

    story.append(cat_table)

    # Build PDF
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)

    return str(filename)


# ============================================================================
# 4. HIGH CONFIDENCE TRANSACTIONS REPORT
# ============================================================================

def generate_high_confidence_report(
    session: Session,
    tax_year_start: datetime,
    tax_year_end: datetime,
    min_confidence: int = 70
) -> str:
    """
    Generate report of high-confidence transactions.

    Shows only transactions with confidence >= threshold for HMRC submission.
    """
    transactions = session.query(Transaction).filter(
        and_(
            Transaction.date >= tax_year_start,
            Transaction.date <= tax_year_end,
            Transaction.confidence_score >= min_confidence
        )
    ).order_by(Transaction.date.desc()).all()

    tax_year_str = f"{tax_year_start.year}/{str(tax_year_end.year)[2:]}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"high_confidence_{tax_year_str.replace('/', '_')}_{timestamp}.csv"

    data = []
    for trans in transactions:
        data.append({
            'Date': trans.date.strftime('%Y-%m-%d'),
            'Merchant': trans.merchant or '',
            'Description': trans.description or '',
            'Amount': f"{trans.amount:.2f}",
            'Type': trans.type.value if trans.type else '',
            'Category': trans.category or '',
            'Confidence': f"{trans.confidence_score}%",
            'Method': trans.categorization_method or ''
        })

    df = pd.DataFrame(data)
    df.to_csv(filename, index=False)

    return str(filename)


# ============================================================================
# 5. REQUIRES REVIEW REPORT
# ============================================================================

def generate_requires_review_report(
    session: Session,
    tax_year_start: datetime,
    tax_year_end: datetime,
    max_confidence: int = 40
) -> str:
    """
    Generate report of transactions requiring manual review.

    Lists all transactions with low confidence for verification.
    """
    transactions = session.query(Transaction).filter(
        and_(
            Transaction.date >= tax_year_start,
            Transaction.date <= tax_year_end,
            or_(
                Transaction.confidence_score < max_confidence,
                Transaction.confidence_score.is_(None),
                Transaction.category.is_(None)
            )
        )
    ).order_by(Transaction.confidence_score.asc()).all()

    tax_year_str = f"{tax_year_start.year}/{str(tax_year_end.year)[2:]}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"requires_review_{tax_year_str.replace('/', '_')}_{timestamp}.csv"

    data = []
    for trans in transactions:
        data.append({
            'Transaction ID': trans.id,
            'Date': trans.date.strftime('%Y-%m-%d'),
            'Merchant': trans.merchant or '',
            'Description': trans.description or '',
            'Amount': f"{trans.amount:.2f}",
            'Type': trans.type.value if trans.type else 'UNCATEGORIZED',
            'Category': trans.category or 'UNCATEGORIZED',
            'Confidence': f"{trans.confidence_score or 0}%",
            'Reason': _get_review_reason(trans)
        })

    df = pd.DataFrame(data)
    df.to_csv(filename, index=False)

    return str(filename)


def _get_review_reason(trans: Transaction) -> str:
    """Determine why transaction requires review."""
    reasons = []

    if not trans.category:
        reasons.append("No category")
    if not trans.confidence_score or trans.confidence_score < 10:
        reasons.append("No confidence score")
    elif trans.confidence_score < 40:
        reasons.append("Low confidence")
    if not trans.type:
        reasons.append("No type")
    if not trans.merchant:
        reasons.append("No merchant")

    return "; ".join(reasons) if reasons else "Review needed"


# ============================================================================
# 6. SA103S COMPATIBLE EXPORT
# ============================================================================

# SA103S expense categories mapping
SA103S_CATEGORIES = {
    17: "Cost of goods bought for resale or materials",
    18: "Car, van and travel expenses",
    19: "Wages, salaries and other staff costs",
    20: "Rent, rates, power and insurance costs",
    21: "Repairs and renewals of property and equipment",
    22: "Phone, fax, stationery and other office costs",
    23: "Advertising and business entertainment costs",
    24: "Interest on bank and other loans",
    25: "Bank, credit card and other financial charges",
    26: "Irrecoverable debts written off",
    27: "Accountancy, legal and other professional fees",
    28: "Depreciation and loss/profit on sale of assets",
    29: "Other business expenses"
}

# Category mapping to SA103S boxes
CATEGORY_TO_SA103S = {
    "Travel": 18,
    "Mileage": 18,
    "Fuel": 18,
    "Parking": 18,
    "Public transport": 18,

    "Office costs": 22,
    "Phone": 22,
    "Internet": 22,
    "Stationery": 22,
    "Postage": 22,

    "Marketing": 23,
    "Advertising": 23,
    "Entertainment": 23,

    "Accountancy": 27,
    "Legal fees": 27,
    "Professional fees": 27,

    "Rent": 20,
    "Insurance": 20,
    "Utilities": 20,

    "Bank charges": 25,
    "Finance charges": 25,

    "Equipment": 21,
    "Repairs": 21,
    "Maintenance": 21,
}


def export_sa103s_format(
    session: Session,
    tax_year_start: datetime,
    tax_year_end: datetime
) -> str:
    """
    Export expenses in HMRC SA103S format.

    Returns CSV with box numbers and totals for Self-Assessment form.
    """
    # Get all expense transactions
    expenses = session.query(Transaction).filter(
        and_(
            Transaction.date >= tax_year_start,
            Transaction.date <= tax_year_end,
            Transaction.type == TransactionType.EXPENSE
        )
    ).all()

    # Initialize box totals
    box_totals = {box: 0.0 for box in SA103S_CATEGORIES.keys()}

    # Categorize expenses into boxes
    for expense in expenses:
        amount = abs(expense.amount)

        # Map category to SA103S box
        box_num = CATEGORY_TO_SA103S.get(expense.category, 29)  # Default to "Other"
        box_totals[box_num] += amount

    # Generate CSV
    tax_year_str = f"{tax_year_start.year}/{str(tax_year_end.year)[2:]}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"sa103s_export_{tax_year_str.replace('/', '_')}_{timestamp}.csv"

    data = []
    for box_num in sorted(box_totals.keys()):
        data.append({
            'Box': box_num,
            'Category': SA103S_CATEGORIES[box_num],
            'Amount': f"{box_totals[box_num]:.2f}"
        })

    df = pd.DataFrame(data)
    df.to_csv(filename, index=False)

    return str(filename)


# ============================================================================
# 7. COMPLETE EXCEL WORKBOOK
# ============================================================================

def generate_excel_workbook(
    session: Session,
    tax_year_start: datetime,
    tax_year_end: datetime
) -> str:
    """
    Generate comprehensive Excel workbook with multiple sheets.

    Sheets:
    1. Summary
    2. Income
    3. Expenses by Category
    4. All Transactions
    5. Audit Trail
    6. Receipts
    """
    tax_year_str = f"{tax_year_start.year}/{str(tax_year_end.year)[2:]}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"complete_workbook_{tax_year_str.replace('/', '_')}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    # Get all data
    transactions = session.query(Transaction).filter(
        and_(
            Transaction.date >= tax_year_start,
            Transaction.date <= tax_year_end
        )
    ).order_by(Transaction.date.desc()).all()

    audit_logs = session.query(AuditLog).filter(
        and_(
            AuditLog.timestamp >= tax_year_start,
            AuditLog.timestamp <= tax_year_end
        )
    ).order_by(AuditLog.timestamp.desc()).all()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_summary_sheet(ws_summary, transactions, tax_year_str)

    # Sheet 2: Income
    ws_income = wb.create_sheet("Income")
    income_trans = [t for t in transactions if t.type == TransactionType.INCOME]
    _create_transactions_sheet(ws_income, income_trans, "Income")

    # Sheet 3: Expenses by Category
    ws_expenses = wb.create_sheet("Expenses by Category")
    expense_trans = [t for t in transactions if t.type == TransactionType.EXPENSE]
    _create_expenses_by_category_sheet(ws_expenses, expense_trans)

    # Sheet 4: All Transactions
    ws_all = wb.create_sheet("All Transactions")
    _create_transactions_sheet(ws_all, transactions, "All Transactions")

    # Sheet 5: Audit Trail
    ws_audit = wb.create_sheet("Audit Trail")
    _format_excel_audit_log_sheet(ws_audit, audit_logs)

    # Sheet 6: Receipts
    ws_receipts = wb.create_sheet("Receipts")
    _create_receipts_sheet(ws_receipts, transactions)

    wb.save(filename)
    return str(filename)


def _create_summary_sheet(ws, transactions, tax_year_str):
    """Create summary sheet in Excel workbook."""
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Tax Summary - {tax_year_str}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])  # Blank row

    # Calculate totals
    income = sum(t.amount for t in transactions if t.type == TransactionType.INCOME)
    expenses = sum(abs(t.amount) for t in transactions if t.type == TransactionType.EXPENSE)
    net = income - expenses

    income_count = sum(1 for t in transactions if t.type == TransactionType.INCOME)
    expense_count = sum(1 for t in transactions if t.type == TransactionType.EXPENSE)

    # Summary data
    ws.append(['Metric', 'Value'])
    ws.append(['Total Income', income])
    ws.append(['Total Expenses', expenses])
    ws.append(['Net Profit/Loss', net])
    ws.append([])
    ws.append(['Income Transactions', income_count])
    ws.append(['Expense Transactions', expense_count])
    ws.append(['Total Transactions', len(transactions)])

    # Format
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    ws['A3'].fill = header_fill
    ws['B3'].fill = header_fill
    ws['A3'].font = Font(bold=True, color="FFFFFF")
    ws['B3'].font = Font(bold=True, color="FFFFFF")

    # Currency formatting
    for row in [4, 5, 6]:
        ws[f'B{row}'].number_format = '£#,##0.00'

    # Number formatting
    for row in [8, 9, 10]:
        ws[f'B{row}'].number_format = '0'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _create_transactions_sheet(ws, transactions, title):
    """Create transactions sheet in Excel workbook."""
    ws.append([title])
    ws.merge_cells(f'A1:{get_column_letter(8)}1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Headers
    headers = ['Date', 'Merchant', 'Description', 'Amount', 'Type', 'Category', 'Confidence', 'Method']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="16a085", end_color="16a085", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for trans in transactions:
        ws.append([
            trans.date.strftime('%d/%m/%Y'),
            trans.merchant or '',
            trans.description or '',
            trans.amount,
            trans.type.value if trans.type else '',
            trans.category or '',
            trans.confidence_score or 0,
            trans.categorization_method or ''
        ])

    # Format amount column
    for row in range(4, len(transactions) + 4):
        ws.cell(row=row, column=4).number_format = '£#,##0.00'
        ws.cell(row=row, column=7).number_format = '0"%"'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18


def _create_expenses_by_category_sheet(ws, expenses):
    """Create expenses by category sheet."""
    ws.append(['Expenses by Category'])
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Group by category
    category_totals = {}
    category_counts = {}

    for expense in expenses:
        cat = expense.category or "Uncategorized"
        category_totals[cat] = category_totals.get(cat, 0) + abs(expense.amount)
        category_counts[cat] = category_counts.get(cat, 0) + 1

    # Headers
    ws.append(['Category', 'Count', 'Total'])

    header_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # Data
    for category in sorted(category_totals.keys()):
        ws.append([
            category,
            category_counts[category],
            category_totals[category]
        ])

    # Format
    start_row = 4
    end_row = start_row + len(category_totals) - 1

    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=2).number_format = '0'
        ws.cell(row=row, column=3).number_format = '£#,##0.00'

    # Total row
    ws.append([])
    total_row = end_row + 2
    ws[f'A{total_row}'] = 'TOTAL'
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'B{total_row}'] = f'=SUM(B{start_row}:B{end_row})'
    ws[f'C{total_row}'] = f'=SUM(C{start_row}:C{end_row})'
    ws[f'C{total_row}'].number_format = '£#,##0.00'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15


def _create_receipts_sheet(ws, transactions):
    """Create receipts sheet."""
    ws.append(['Receipts'])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Headers
    headers = ['Date', 'Merchant', 'Amount', 'Category', 'Receipt File']
    ws.append(headers)

    header_fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # Data
    for trans in transactions:
        if trans.receipts:
            for receipt in trans.receipts:
                ws.append([
                    trans.date.strftime('%d/%m/%Y'),
                    trans.merchant or '',
                    abs(trans.amount),
                    trans.category or '',
                    Path(receipt.file_path).name if receipt.file_path else ''
                ])

    # Format
    for row in range(4, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = '£#,##0.00'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35


def _format_excel_summary_sheet(
    ws, transactions, income_count, expense_count,
    total_changes, audit_entries, tax_year_str
):
    """Format Excel summary sheet for audit trail."""
    ws.append([f'Audit Trail Summary - {tax_year_str}'])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(['Metric', 'Count'])
    ws.append(['Total Transactions', len(transactions)])
    ws.append(['Income Records', income_count])
    ws.append(['Expense Records', expense_count])
    ws.append(['Total Changes Logged', total_changes])
    ws.append(['Audit Log Entries', audit_entries])

    # Format
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    ws['A3'].fill = header_fill
    ws['B3'].fill = header_fill
    ws['A3'].font = Font(bold=True, color="FFFFFF")
    ws['B3'].font = Font(bold=True, color="FFFFFF")

    for row in range(4, 10):
        ws[f'B{row}'].number_format = '0'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def _format_excel_audit_log_sheet(ws, audit_logs):
    """Format Excel audit log sheet."""
    # Headers
    headers = ['Timestamp', 'Transaction ID', 'Action', 'Before Value', 'After Value', 'Change Summary']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="16a085", end_color="16a085", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for log in audit_logs:
        ws.append([
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.transaction_id or '',
            log.action,
            log.before_value or '',
            log.after_value or '',
            log.change_summary or ''
        ])

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40


# ============================================================================
# PDF UTILITIES
# ============================================================================

def _add_page_number(canvas_obj, doc):
    """Add page number to PDF footer."""
    page_num = canvas_obj.getPageNumber()
    text = f"Page {page_num}"
    canvas_obj.setFont('Helvetica', 9)
    canvas_obj.setFillColor(colors.grey)
    canvas_obj.drawRightString(7.5*inch, 0.5*inch, text)

    # Add generation timestamp
    timestamp = datetime.now().strftime('%d %B %Y at %H:%M')
    canvas_obj.drawString(0.75*inch, 0.5*inch, f"Generated: {timestamp}")


# ============================================================================
# REPORT ARCHIVE
# ============================================================================

def save_report_to_archive(report_data: bytes, report_type: str, tax_year: str) -> str:
    """
    Save generated report to reports archive directory.

    Args:
        report_data: Binary report data
        report_type: Type of report (e.g., 'audit_trail', 'receipt_summary')
        tax_year: Tax year string (e.g., '2024/25')

    Returns:
        Path to saved file
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"{report_type}_{tax_year.replace('/', '_')}_{timestamp}.pdf"

    with open(filename, 'wb') as f:
        f.write(report_data)

    return str(filename)


def list_archived_reports() -> List[Dict[str, Any]]:
    """
    List all archived reports with metadata.

    Returns:
        List of report metadata dictionaries
    """
    reports = []

    for file_path in REPORTS_DIR.glob("*.pdf"):
        stat = file_path.stat()
        reports.append({
            'filename': file_path.name,
            'path': str(file_path),
            'size_mb': stat.st_size / (1024 * 1024),
            'created': datetime.fromtimestamp(stat.st_ctime),
            'type': file_path.stem.rsplit('_', 2)[0]
        })

    # Sort by creation date, newest first
    reports.sort(key=lambda x: x['created'], reverse=True)

    return reports


# ============================================================================
# REPORT GENERATOR UI
# ============================================================================

def render_report_generator_ui(session: Session):
    """
    Render report generator UI component.

    Provides interface for generating various compliance reports.
    """
    print("\n" + "="*60)
    print("GENERATE COMPLIANCE REPORTS".center(60))
    print("="*60 + "\n")

    print("Select Report Type:")
    print("  1. Audit Trail Report (PDF)")
    print("  2. Receipt Summary (PDF)")
    print("  3. Categorization Report (PDF)")
    print("  4. High Confidence Transactions (CSV)")
    print("  5. Requires Review Report (CSV)")
    print("  6. SA103S Export (CSV)")
    print("  7. Complete Workbook (Excel)")
    print("  8. View Previously Generated Reports")
    print("  0. Back")

    choice = input("\nEnter choice (0-8): ").strip()

    if choice == '0':
        return

    if choice == '8':
        _display_archived_reports()
        return

    # Get tax year
    current_tax_year = get_current_tax_year()
    print(f"\nCurrent Tax Year: {current_tax_year}")
    tax_year = input(f"Enter tax year [{current_tax_year}]: ").strip() or current_tax_year

    try:
        start_date, end_date = get_tax_year_dates(tax_year)
    except Exception as e:
        print(f"\nError: Invalid tax year format. Use YYYY/YY (e.g., 2024/25)")
        return

    print(f"\nGenerating report for {tax_year}...")
    print(f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}")

    try:
        if choice == '1':
            filepath = generate_audit_trail_report(session, start_date, end_date, 'PDF')
            print(f"\nAudit trail report generated: {filepath}")

        elif choice == '2':
            filepath = generate_receipt_summary(session, start_date, end_date, 'PDF')
            print(f"\nReceipt summary report generated: {filepath}")

        elif choice == '3':
            filepath = generate_categorization_report(session, start_date, end_date)
            print(f"\nCategorization report generated: {filepath}")

        elif choice == '4':
            filepath = generate_high_confidence_report(session, start_date, end_date, 70)
            print(f"\nHigh confidence report generated: {filepath}")

        elif choice == '5':
            filepath = generate_requires_review_report(session, start_date, end_date, 40)
            print(f"\nRequires review report generated: {filepath}")

        elif choice == '6':
            filepath = export_sa103s_format(session, start_date, end_date)
            print(f"\nSA103S export generated: {filepath}")

        elif choice == '7':
            filepath = generate_excel_workbook(session, start_date, end_date)
            print(f"\nComplete workbook generated: {filepath}")

        else:
            print("\nInvalid choice.")
            return

        # Show file info
        file_size = os.path.getsize(filepath) / (1024 * 1024)
        print(f"File size: {file_size:.2f} MB")
        print(f"\nReport saved to: {filepath}")

    except Exception as e:
        print(f"\nError generating report: {e}")
        import traceback
        traceback.print_exc()


def _display_archived_reports():
    """Display list of previously generated reports."""
    reports = list_archived_reports()

    if not reports:
        print("\nNo reports found in archive.")
        return

    print("\n" + "="*80)
    print("PREVIOUSLY GENERATED REPORTS".center(80))
    print("="*80 + "\n")

    for idx, report in enumerate(reports[:20], 1):  # Show last 20
        print(f"{idx}. {report['filename']}")
        print(f"   Type: {report['type']}")
        print(f"   Size: {report['size_mb']:.2f} MB")
        print(f"   Created: {report['created'].strftime('%d %B %Y at %H:%M')}")
        print(f"   Path: {report['path']}")
        print()

    if len(reports) > 20:
        print(f"... and {len(reports) - 20} more reports")


# ============================================================================
# CONVENIENCE WRAPPER FUNCTIONS
# ============================================================================

def generate_sa103s_export(session: Session, tax_year: str) -> str:
    """
    Convenience wrapper for export_sa103s_format()
    Generate HMRC SA103S format export

    Args:
        session: SQLAlchemy session
        tax_year: Tax year string (e.g., "2023/24")

    Returns:
        Path to generated CSV file
    """
    return export_sa103s_format(session, tax_year)


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    from database import SessionLocal

    session = SessionLocal()
    try:
        render_report_generator_ui(session)
    finally:
        session.close()
