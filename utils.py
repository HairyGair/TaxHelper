"""
Utility functions for UK Self Assessment Tax Helper
Handles CSV parsing, date parsing, duplicate detection, Excel export, and rules application
"""

import pandas as pd
import re
from datetime import datetime
from dateutil import parser as date_parser
from typing import List, Dict, Tuple, Optional
import io

# Import smart categorization modules
try:
    from merchant_database import get_categorization_confidence
    from pattern_analyzer import analyze_transactions, merge_confidence_scores
    SMART_CATEGORIZATION_AVAILABLE = True
except ImportError:
    SMART_CATEGORIZATION_AVAILABLE = False


def parse_uk_date(date_string: str) -> Optional[datetime]:
    """
    Parse UK date formats flexibly
    Handles DD/MM/YYYY, DD-MM-YYYY, YYYY-MM-DD, etc.

    Security: Validates date ranges to prevent invalid data
    """
    if pd.isna(date_string) or not date_string:
        return None

    # Try common UK formats first
    uk_formats = [
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%d.%m.%Y',
        '%d/%m/%y',
        '%d-%m-%y',
        '%d-%b-%y',  # NatWest format: 04-Apr-25
        '%d-%B-%y',  # Full month name
    ]

    parsed_date = None
    for fmt in uk_formats:
        try:
            parsed_date = datetime.strptime(str(date_string).strip(), fmt)
            break
        except ValueError:
            continue

    # Fall back to dateutil parser (handles ISO and other formats)
    if not parsed_date:
        try:
            # dayfirst=True assumes DD/MM/YYYY for ambiguous dates
            parsed_date = date_parser.parse(str(date_string).strip(), dayfirst=True)
        except:
            return None

    # Security: Validate date range
    # Reject future dates and dates older than 100 years
    if parsed_date:
        today = datetime.now()
        min_date = datetime(today.year - 100, 1, 1)
        max_date = today + timedelta(days=1)  # Allow today + 1 day tolerance

        if parsed_date < min_date or parsed_date > max_date:
            return None  # Invalid date range

    return parsed_date


def parse_currency(value: str) -> float:
    """
    Parse currency string to float
    Handles £, commas, negative values

    Security: Validates amount ranges to prevent invalid data
    """
    if pd.isna(value) or value == '' or value is None:
        return 0.0

    # Convert to string and clean
    value_str = str(value).strip()

    # Remove currency symbols and whitespace
    value_str = value_str.replace('£', '').replace('$', '').replace(',', '').strip()

    # Handle empty strings after cleaning
    if not value_str:
        return 0.0

    try:
        amount = float(value_str)

        # Security: Validate reasonable amount range
        # Reject amounts outside reasonable bounds (-1 billion to +1 billion)
        MAX_AMOUNT = 1_000_000_000.0
        if abs(amount) > MAX_AMOUNT:
            return 0.0

        return amount
    except ValueError:
        return 0.0


def format_currency(value: float) -> str:
    """
    Format float as UK currency string
    """
    return f"£{value:,.2f}"


def detect_duplicates(df: pd.DataFrame, session, Transaction) -> List[int]:
    """
    Detect duplicate transactions already in database using batch query
    Returns list of row indices that are duplicates
    """
    from sqlalchemy import and_, tuple_

    duplicates = []

    if df.empty:
        return duplicates

    # Build a single batch query using filter().in_() - avoids N+1 query problem
    # Create tuples of (date, description, paid_in, paid_out) for all rows
    transaction_tuples = []
    for idx, row in df.iterrows():
        date_val = row.get('date')
        desc_val = row.get('description', '')
        paid_in = row.get('paid_in', 0.0)
        paid_out = row.get('paid_out', 0.0)
        transaction_tuples.append((date_val, desc_val, paid_in, paid_out))

    # Query all potential duplicates in a single database call
    existing_transactions = session.query(
        Transaction.date,
        Transaction.description,
        Transaction.paid_in,
        Transaction.paid_out
    ).filter(
        tuple_(
            Transaction.date,
            Transaction.description,
            Transaction.paid_in,
            Transaction.paid_out
        ).in_(transaction_tuples)
    ).all()

    # Convert to set for O(1) lookup
    existing_set = set(existing_transactions)

    # Check which rows are duplicates
    for idx, row in df.iterrows():
        date_val = row.get('date')
        desc_val = row.get('description', '')
        paid_in = row.get('paid_in', 0.0)
        paid_out = row.get('paid_out', 0.0)

        if (date_val, desc_val, paid_in, paid_out) in existing_set:
            duplicates.append(idx)

    return duplicates


def apply_rules(description: str, paid_in: float, paid_out: float, rules: List) -> Tuple[Optional[str], Optional[str], bool]:
    """
    Apply categorization rules to a transaction
    Returns (guessed_type, guessed_category, is_personal)
    """
    # Sort rules by priority (lower = higher priority)
    sorted_rules = sorted([r for r in rules if r.enabled], key=lambda x: x.priority)

    for rule in sorted_rules:
        match = False

        if rule.match_mode == 'Contains':
            match = rule.text_to_match.lower() in description.lower()
        elif rule.match_mode == 'Equals':
            match = rule.text_to_match.lower() == description.lower()
        elif rule.match_mode == 'Regex':
            try:
                match = bool(re.search(rule.text_to_match, description, re.IGNORECASE))
            except re.error:
                continue

        if match:
            guessed_type = rule.map_to
            guessed_category = None
            is_personal = rule.is_personal if hasattr(rule, 'is_personal') else False

            if rule.map_to == 'Income':
                guessed_category = rule.income_type
            elif rule.map_to == 'Expense':
                guessed_category = rule.expense_category

            return guessed_type, guessed_category, is_personal

    # Default: if paid_in > 0, guess Income; if paid_out > 0, guess Expense
    # Defaults to business (is_personal=False)
    if paid_in > 0:
        return 'Income', 'Self-employment', False
    elif paid_out > 0:
        return 'Expense', 'Other business expenses', False

    return None, None, False


def apply_smart_categorization(session, transactions: List) -> None:
    """
    Apply intelligent categorization using merchant database and pattern analysis
    Updates transactions with confidence scores and improved categorization

    Args:
        session: SQLAlchemy session
        transactions: List of Transaction model instances (already saved to DB)
    """
    if not SMART_CATEGORIZATION_AVAILABLE:
        print("Warning: Smart categorization modules not available, using basic rules only")
        return

    if not transactions:
        return

    print(f"Running smart categorization on {len(transactions)} transactions...")

    # Step 1: Run pattern analysis on all transactions
    pattern_results = analyze_transactions(session, transactions)

    # Step 2: For each transaction, combine merchant and pattern data
    for txn in transactions:
        # Get merchant confidence
        is_personal_merchant, category_merchant, confidence_merchant, merchant_name = \
            get_categorization_confidence(txn.description)

        # Get pattern analysis result
        pattern_result = pattern_results.get(txn.id)

        # Start with defaults
        final_type = txn.guessed_type
        final_category = txn.guessed_category
        final_is_personal = txn.is_personal
        final_confidence = 0
        requires_review = False

        # Combine both sources of information
        if pattern_result and pattern_result.pattern_confidence > 0:
            # We have pattern analysis
            pattern_conf = pattern_result.pattern_confidence
            pattern_personal = pattern_result.is_personal
            pattern_type = pattern_result.suggested_type
            pattern_category = pattern_result.suggested_category

            # Merge with merchant data
            combined_conf, combined_personal = merge_confidence_scores(
                pattern_confidence=pattern_conf,
                merchant_confidence=confidence_merchant,
                pattern_type=pattern_result.primary_pattern.pattern_type if pattern_result.primary_pattern else None,
                pattern_personal=pattern_personal,
                merchant_personal=is_personal_merchant
            )

            final_confidence = combined_conf
            final_is_personal = combined_personal

            # Use pattern suggestion if high confidence
            if pattern_conf >= 70 and pattern_type:
                final_type = pattern_type
                if pattern_category:
                    final_category = pattern_category
            # Otherwise use merchant suggestion if available
            elif confidence_merchant >= 70:
                if txn.paid_in > 0:
                    final_type = "Income"
                elif txn.paid_out > 0:
                    final_type = "Expense"
                final_category = category_merchant

            # Store pattern metadata
            if pattern_result.primary_pattern:
                txn.pattern_type = pattern_result.primary_pattern.pattern_type.value
                txn.pattern_group_id = pattern_result.primary_pattern.metadata.get('group_id', '')
                txn.pattern_metadata = pattern_result.primary_pattern.metadata

            requires_review = pattern_result.requires_review

        elif confidence_merchant > 0:
            # Only merchant data available
            final_confidence = confidence_merchant
            final_is_personal = is_personal_merchant

            if confidence_merchant >= 70:
                if txn.paid_in > 0:
                    final_type = "Income"
                elif txn.paid_out > 0:
                    final_type = "Expense"
                final_category = category_merchant

        # Update transaction with smart categorization
        txn.guessed_type = final_type
        txn.guessed_category = final_category
        txn.is_personal = final_is_personal
        txn.confidence_score = final_confidence
        txn.merchant_confidence = confidence_merchant
        txn.pattern_confidence = pattern_result.pattern_confidence if pattern_result else 0
        txn.requires_review = requires_review

    # Commit all updates
    session.commit()
    print(f"✓ Smart categorization complete")

    # Print summary
    high_conf = sum(1 for t in transactions if t.confidence_score >= 70)
    med_conf = sum(1 for t in transactions if 40 <= t.confidence_score < 70)
    low_conf = sum(1 for t in transactions if t.confidence_score < 40)
    personal = sum(1 for t in transactions if t.is_personal)
    business = sum(1 for t in transactions if not t.is_personal)

    print(f"  High confidence (≥70): {high_conf} ({high_conf/len(transactions)*100:.1f}%)")
    print(f"  Medium confidence (40-69): {med_conf} ({med_conf/len(transactions)*100:.1f}%)")
    print(f"  Low confidence (<40): {low_conf} ({low_conf/len(transactions)*100:.1f}%)")
    print(f"  Personal: {personal} ({personal/len(transactions)*100:.1f}%)")
    print(f"  Business: {business} ({business/len(transactions)*100:.1f}%)")


def parse_csv(
    file_content: bytes,
    column_mappings: Dict[str, str],
    session,
    rules: List,
    Transaction
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Parse CSV file with flexible column mapping
    Apply rules and detect duplicates
    Returns (dataframe, list of errors)

    Security: Validates file size, row count, and field lengths
    """
    errors = []

    # Security: Validate file size (max 50MB)
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    if len(file_content) > MAX_FILE_SIZE:
        errors.append(f"File too large. Maximum size is {MAX_FILE_SIZE / 1024 / 1024:.0f}MB")
        return None, errors

    # Security: Validate file is not empty
    if len(file_content) == 0:
        errors.append("File is empty")
        return None, errors

    try:
        # Try to read CSV with different encodings
        try:
            df = pd.read_csv(io.BytesIO(file_content), encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(io.BytesIO(file_content), encoding='latin-1')
            except Exception as e:
                errors.append(f"Unable to parse CSV file. Please check file encoding.")
                return None, errors

        # Security: Validate row count (max 100,000 rows)
        MAX_ROWS = 100_000
        if len(df) > MAX_ROWS:
            errors.append(f"Too many rows in CSV. Maximum is {MAX_ROWS:,} rows")
            return None, errors

        # Security: Validate column count (max 50 columns)
        MAX_COLUMNS = 50
        if len(df.columns) > MAX_COLUMNS:
            errors.append(f"Too many columns in CSV. Maximum is {MAX_COLUMNS} columns")
            return None, errors

        # Check if using single Value column or separate Paid in/Paid out columns
        use_value_column = False
        if column_mappings.get('column_value') and column_mappings['column_value'] in df.columns:
            use_value_column = True
        elif (not column_mappings.get('column_paid_in') or column_mappings['column_paid_in'] not in df.columns) and \
             (not column_mappings.get('column_paid_out') or column_mappings['column_paid_out'] not in df.columns):
            # Try to find Value column if paid_in/paid_out are missing
            if 'Value' in df.columns:
                column_mappings['column_value'] = 'Value'
                use_value_column = True

        # Validate required columns exist
        required_columns = ['column_date', 'column_description']
        if use_value_column:
            if column_mappings.get('column_value') not in df.columns:
                errors.append(f"Required column '{column_mappings.get('column_value')}' not found in CSV")
        else:
            if column_mappings.get('column_paid_in') not in df.columns:
                errors.append(f"Required column '{column_mappings.get('column_paid_in')}' not found in CSV")
            if column_mappings.get('column_paid_out') not in df.columns:
                errors.append(f"Required column '{column_mappings.get('column_paid_out')}' not found in CSV")

        for col_key in required_columns:
            expected_col = column_mappings.get(col_key)
            if not expected_col or expected_col not in df.columns:
                errors.append(f"Required column '{expected_col}' not found in CSV")

        if errors:
            return None, errors

        # Map columns to standard names
        rename_map = {
            column_mappings['column_date']: 'date',
            column_mappings['column_description']: 'description',
        }

        if use_value_column:
            rename_map[column_mappings['column_value']] = 'value'
        else:
            rename_map[column_mappings['column_paid_in']] = 'paid_in'
            rename_map[column_mappings['column_paid_out']] = 'paid_out'

        if column_mappings.get('column_type') and column_mappings['column_type'] in df.columns:
            rename_map[column_mappings['column_type']] = 'type'
        if column_mappings.get('column_balance') and column_mappings['column_balance'] in df.columns:
            rename_map[column_mappings['column_balance']] = 'balance'

        df = df.rename(columns=rename_map)

        # Security: Validate and sanitize description field length
        MAX_DESCRIPTION_LENGTH = 500
        if 'description' in df.columns:
            df['description'] = df['description'].apply(
                lambda x: str(x)[:MAX_DESCRIPTION_LENGTH] if pd.notna(x) else ''
            )

        # Parse dates
        df['date'] = df['date'].apply(parse_uk_date)

        # Remove rows with invalid dates
        invalid_dates = df['date'].isna().sum()
        if invalid_dates > 0:
            errors.append(f"Warning: {invalid_dates} rows with invalid dates were removed")
            df = df[df['date'].notna()]

        # Security: Validate no rows remain after date filtering
        if len(df) == 0:
            errors.append("No valid transactions found after filtering invalid dates")
            return None, errors

        # Parse currency values and split Value column if needed
        if use_value_column:
            df['value'] = df['value'].apply(parse_currency)
            # Split into paid_in and paid_out based on sign
            df['paid_in'] = df['value'].apply(lambda x: x if x > 0 else 0.0)
            df['paid_out'] = df['value'].apply(lambda x: abs(x) if x < 0 else 0.0)
        else:
            df['paid_in'] = df['paid_in'].apply(parse_currency)
            df['paid_out'] = df['paid_out'].apply(parse_currency)

        if 'balance' in df.columns:
            df['balance'] = df['balance'].apply(parse_currency)

        # Apply rules to guess categories and personal/business flag
        df['guessed_type'] = None
        df['guessed_category'] = None
        df['is_personal'] = False

        for idx, row in df.iterrows():
            guessed_type, guessed_category, is_personal = apply_rules(
                row['description'],
                row['paid_in'],
                row['paid_out'],
                rules
            )
            df.at[idx, 'guessed_type'] = guessed_type
            df.at[idx, 'guessed_category'] = guessed_category
            df.at[idx, 'is_personal'] = is_personal

        # Detect duplicates (with error handling)
        try:
            duplicate_indices = detect_duplicates(df, session, Transaction)
            if duplicate_indices:
                errors.append(f"Warning: {len(duplicate_indices)} duplicate transactions detected and will be skipped")
                df = df.drop(duplicate_indices)
        except Exception as e:
            # Non-critical error - log but continue
            errors.append(f"Warning: Could not check for duplicates")

        # Security: Validate final dataframe has valid data
        if len(df) == 0:
            errors.append("No valid transactions to import after processing")
            return None, errors

        # Add default values
        df['reviewed'] = False
        df['notes'] = ''

        return df, errors

    except pd.errors.ParserError as e:
        # Security: Don't expose detailed parser errors
        errors.append("CSV file format is invalid. Please check the file structure.")
        return None, errors
    except MemoryError:
        errors.append("File is too large to process. Please reduce file size.")
        return None, errors
    except Exception as e:
        # Security: Generic error message to avoid information disclosure
        errors.append("Unable to parse CSV file. Please check file format and try again.")
        # Log actual error for debugging
        import logging
        logging.error(f"CSV parse error: {str(e)}")
        return None, errors


def export_to_excel(file_path: str, session, models_dict: Dict, settings: Dict[str, str]):
    """
    Export all data to Excel workbook
    Creates sheets for each data type plus summary
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Extract models
    Transaction = models_dict['Transaction']
    Income = models_dict['Income']
    Expense = models_dict['Expense']
    Mileage = models_dict['Mileage']
    Donation = models_dict['Donation']
    Rule = models_dict['Rule']
    Setting = models_dict['Setting']

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 1. Profile Sheet
    ws_profile = wb.create_sheet("Profile")
    ws_profile.append(["Setting", "Value"])
    for cell in ws_profile[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for key, value in settings.items():
        ws_profile.append([key, value])

    # 2. Income Sheet
    income_records = session.query(Income).order_by(Income.date.desc()).all()
    ws_income = wb.create_sheet("Income")
    ws_income.append(["Date", "Source", "Description", "Amount (Gross)", "Tax Deducted", "Income Type", "Notes"])

    for cell in ws_income[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for record in income_records:
        ws_income.append([
            record.date.strftime('%d/%m/%Y') if record.date else '',
            record.source,
            record.description,
            record.amount_gross,
            record.tax_deducted,
            record.income_type,
            record.notes
        ])

    # 3. Expenses Sheet
    expense_records = session.query(Expense).order_by(Expense.date.desc()).all()
    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses.append(["Date", "Supplier", "Description", "Category", "Amount", "Receipt Link", "Notes"])

    for cell in ws_expenses[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for record in expense_records:
        ws_expenses.append([
            record.date.strftime('%d/%m/%Y') if record.date else '',
            record.supplier,
            record.description,
            record.category,
            record.amount,
            record.receipt_link,
            record.notes
        ])

    # 4. Mileage Sheet
    mileage_records = session.query(Mileage).order_by(Mileage.date.desc()).all()
    ws_mileage = wb.create_sheet("Mileage")
    ws_mileage.append(["Date", "Purpose", "From", "To", "Miles", "Rate/Mile", "Allowable Amount", "Notes"])

    for cell in ws_mileage[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for record in mileage_records:
        ws_mileage.append([
            record.date.strftime('%d/%m/%Y') if record.date else '',
            record.purpose,
            record.from_location,
            record.to_location,
            record.miles,
            record.rate_per_mile,
            record.allowable_amount,
            record.notes
        ])

    # 5. Donations Sheet
    donation_records = session.query(Donation).order_by(Donation.date.desc()).all()
    ws_donations = wb.create_sheet("Donations")
    ws_donations.append(["Date", "Charity", "Amount Paid", "Gift Aid", "Notes"])

    for cell in ws_donations[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for record in donation_records:
        ws_donations.append([
            record.date.strftime('%d/%m/%Y') if record.date else '',
            record.charity,
            record.amount_paid,
            "Yes" if record.gift_aid else "No",
            record.notes
        ])

    # 6. Summary Sheet (HMRC totals)
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append([f"HMRC Self Assessment Summary for Tax Year {settings.get('tax_year', 'N/A')}"])
    ws_summary.append([])

    # Calculate totals
    from sqlalchemy import func

    # Employment income
    employment_total = session.query(func.sum(Income.amount_gross)).filter(
        Income.income_type == 'Employment'
    ).scalar() or 0.0
    employment_tax = session.query(func.sum(Income.tax_deducted)).filter(
        Income.income_type == 'Employment'
    ).scalar() or 0.0

    # Self-employment income
    self_employment_total = session.query(func.sum(Income.amount_gross)).filter(
        Income.income_type == 'Self-employment'
    ).scalar() or 0.0

    # Interest
    interest_total = session.query(func.sum(Income.amount_gross)).filter(
        Income.income_type == 'Interest'
    ).scalar() or 0.0

    # Dividends
    dividends_total = session.query(func.sum(Income.amount_gross)).filter(
        Income.income_type == 'Dividends'
    ).scalar() or 0.0

    # Expenses
    expenses_total = session.query(func.sum(Expense.amount)).scalar() or 0.0

    # Mileage
    mileage_total = session.query(func.sum(Mileage.allowable_amount)).scalar() or 0.0

    # Total allowable expenses
    total_allowable = expenses_total + mileage_total

    # Net profit
    net_profit = self_employment_total - total_allowable

    # Donations
    donations_total = session.query(func.sum(Donation.amount_paid)).filter(
        Donation.gift_aid == True
    ).scalar() or 0.0

    ws_summary.append(["EMPLOYMENT (if applicable)"])
    ws_summary.append(["Box 1 - Pay from employment", employment_total])
    ws_summary.append(["Box 2 - UK tax deducted", employment_tax])
    ws_summary.append([])

    ws_summary.append(["SELF-EMPLOYMENT (SA103S - Short Form)"])
    ws_summary.append(["Box 15 - Turnover", self_employment_total])
    ws_summary.append(["Box 31 - Total allowable expenses", total_allowable])
    ws_summary.append(["Box 32 - Net profit", net_profit])
    ws_summary.append([])

    ws_summary.append(["SAVINGS INTEREST"])
    ws_summary.append(["Box 1 - Interest (gross)", interest_total])
    ws_summary.append([])

    ws_summary.append(["DIVIDENDS"])
    ws_summary.append(["Box 1 - Dividends (gross)", dividends_total])
    ws_summary.append([])

    ws_summary.append(["GIFT AID"])
    ws_summary.append(["Donations paid", donations_total])
    ws_summary.append([])

    ws_summary.append(["Breakdown of Allowable Expenses:"])
    ws_summary.append(["Total Expenses", expenses_total])
    ws_summary.append(["Total Mileage Allowance", mileage_total])

    # 7. Rules Sheet
    rule_records = session.query(Rule).order_by(Rule.priority).all()
    ws_rules = wb.create_sheet("Rules")
    ws_rules.append(["Match Mode", "Text to Match", "Map To", "Income Type", "Expense Category", "Priority", "Enabled", "Notes"])

    for cell in ws_rules[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for rule in rule_records:
        ws_rules.append([
            rule.match_mode,
            rule.text_to_match,
            rule.map_to,
            rule.income_type or '',
            rule.expense_category or '',
            rule.priority,
            "Yes" if rule.enabled else "No",
            rule.notes
        ])

    # 8. Archived Inbox Sheet (all reviewed transactions)
    transaction_records = session.query(Transaction).filter(
        Transaction.reviewed == True
    ).order_by(Transaction.date.desc()).all()

    ws_inbox = wb.create_sheet("Archived_Inbox")
    ws_inbox.append(["Date", "Type", "Description", "Paid Out", "Paid In", "Balance", "Guessed Type", "Guessed Category", "Personal", "Notes"])

    for cell in ws_inbox[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for record in transaction_records:
        ws_inbox.append([
            record.date.strftime('%d/%m/%Y') if record.date else '',
            record.type,
            record.description,
            record.paid_out,
            record.paid_in,
            record.balance,
            record.guessed_type,
            record.guessed_category,
            "Yes" if record.is_personal else "No",
            record.notes
        ])

    # Adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)


def get_tax_year_dates(tax_year_str: str) -> Tuple[datetime, datetime]:
    """
    Convert tax year string like "2024/25" to start and end dates
    UK tax year runs 6 April to 5 April
    """
    start_year = int(tax_year_str.split('/')[0])
    end_year = start_year + 1

    start_date = datetime(start_year, 4, 6)
    end_date = datetime(end_year, 4, 5)

    return start_date, end_date


def calculate_mileage_allowance(miles: float, cumulative_miles: float = 0) -> Tuple[float, float]:
    """
    Calculate mileage allowance based on HMRC rates
    First 10,000 miles: 45p/mile
    After 10,000 miles: 25p/mile
    Returns (allowable_amount, rate_used)
    """
    remaining_at_high_rate = max(0, 10000 - cumulative_miles)

    if miles <= remaining_at_high_rate:
        # All miles at high rate
        return miles * 0.45, 0.45
    else:
        # Some at high rate, some at low rate
        high_rate_miles = remaining_at_high_rate
        low_rate_miles = miles - remaining_at_high_rate
        total = (high_rate_miles * 0.45) + (low_rate_miles * 0.25)
        avg_rate = total / miles if miles > 0 else 0.45
        return total, avg_rate
